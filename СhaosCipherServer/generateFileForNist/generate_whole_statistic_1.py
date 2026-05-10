import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "summary_table.xlsx"
SYSTEMS = [
    ( r"C:\Users\sasha\Desktop\Диплом 25\ChaosCipher\DiplomAnalys\Експеремент Nist Sts\Фінальні результати обрані за 050426\SHA\ВанДерПоля\ShaОдинSeed\finalAnalysisReport.txt"),
    (r"C:\Users\sasha\Desktop\Диплом 25\ChaosCipher\DiplomAnalys\Експеремент Nist Sts\Фінальні результати обрані за 050426\SHA\ВанДерПоля\ChaОдинSeed\finalAnalysisReport.txt"),
    (r"C:\Users\sasha\Desktop\Диплом 25\ChaosCipher\DiplomAnalys\Експеремент Nist Sts\Фінальні результати обрані за 050426\SHA\ВанДерПоля\ChaДваSeed\finalAnalysisReport.txt"),
]

COL_PASS_PV   = 14   # PASS P_V?
COL_PASS_PR   = 15   # PASS PR?

def read_nist_excel(filepath: str) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    total        = 0
    passed_pv    = 0   # Uniformity (P-VALUE)
    passed_pr    = 0   # Proportion
    passed_both  = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        pv_cell = row[COL_PASS_PV - 1]   # "PASS" / "FAIL"
        pr_cell = row[COL_PASS_PR - 1]

        is_pv   = str(pv_cell).strip().upper() == "PASS"
        is_pr   = str(pr_cell).strip().upper() == "PASS"

        total       += 1
        passed_pv   += int(is_pv)
        passed_pr   += int(is_pr)
        passed_both += int(is_pv and is_pr)

    score = (passed_pv + passed_pr) / (2 * total) if total else 0

    return {
        "total":       total,
        "passed_pv":   passed_pv,    # Uniformity
        "passed_pr":   passed_pr,    # Proportion
        "passed_both": passed_both,
        "score":       round(score, 3),
    }

# ── Собираем данные ───────────────────────────────────────────────────────────
results = []
for name, path in SYSTEMS:
    print(f"Читаю: {name} -> {path}")
    data = read_nist_excel(path)
    results.append((name, data))
    print(f"  Total={data['total']}  PV={data['passed_pv']}  "
          f"PR={data['passed_pr']}  Both={data['passed_both']}  "
          f"Score={data['score']}")

# ── Стили ─────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
BEST_FILL    = PatternFill("solid", start_color="E2EFDA")   # лучший результат
WORST_FILL   = PatternFill("solid", start_color="FFDDC1")   # худший результат
ALT_FILL     = PatternFill("solid", start_color="D6E4F0")
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF")

h_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
d_font  = Font(name="Arial", size=11)
b_font  = Font(name="Arial", size=11, bold=True)
center  = Alignment(horizontal="center", vertical="center")
left    = Alignment(horizontal="left",   vertical="center")
thin    = Side(style="thin", color="BFBFBF")
brd     = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Создаём Excel ─────────────────────────────────────────────────────────────
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "NIST Summary"

headers    = ["Generator", "Passed Proportion", "Passed Uniformity",
              "Passed Both", "Score"]
col_widths = [20, 20, 20, 18, 10]

# Заголовок
for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    c = ws_out.cell(row=1, column=col, value=h)
    c.font      = h_font
    c.fill      = HEADER_FILL
    c.alignment = center
    c.border    = brd
    ws_out.column_dimensions[get_column_letter(col)].width = w
ws_out.row_dimensions[1].height = 24

# Определяем лучший/худший Score для подсветки
scores = [d["score"] for _, d in results]
best_score  = max(scores)
worst_score = min(scores)

# Строки данных
for i, (name, d) in enumerate(results, 1):
    rn = i + 1
    t  = d["total"]

    bg = WHITE_FILL if i % 2 else ALT_FILL
    if d["score"] == best_score:
        bg = BEST_FILL
    elif d["score"] == worst_score:
        bg = WORST_FILL

    vals = [
        name,
        f"{d['passed_pr']}/{t}",
        f"{d['passed_pv']}/{t}",
        f"{d['passed_both']}/{t}",
        d["score"],
    ]
    aligns = [left, center, center, center, center]

    for col, (val, aln) in enumerate(zip(vals, aligns), 1):
        c = ws_out.cell(row=rn, column=col, value=val)
        c.font      = b_font if col == 1 else d_font
        c.fill      = bg
        c.alignment = aln
        c.border    = brd

    # Score — форматируем как число
    score_cell = ws_out.cell(rn, 5, value=d["score"])
    score_cell.number_format = "0.000"
    score_cell.font = Font(
        name="Arial", size=11, bold=True,
        color="375623" if d["score"] >= 0.97 else
              "9C0006" if d["score"] < 0.95 else "7F6000"
    )
    score_cell.fill      = bg
    score_cell.alignment = center
    score_cell.border    = brd

    ws_out.row_dimensions[rn].height = 20

wb_out.save(OUTPUT_FILE)
print(f"\nГотово! Сохранено → {OUTPUT_FILE}")