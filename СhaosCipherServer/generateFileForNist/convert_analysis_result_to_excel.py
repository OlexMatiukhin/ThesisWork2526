import re
import openpyxl
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


#
# SYSTEMS = [
#     (r"C:\Users\sasha\Desktop\Диплом 25\ChaosCipher\DiplomAnalys\Експеремент Nist Sts\Фінальні результати обрані за 050426\SHA\Лоренца\ShaОдинSeed\finalAnalysisReportLorenShaOS.txt"),
#     (r"C:\Users\sasha\Desktop\Диплом 25\ChaosCipher\DiplomAnalys\Експеремент Nist Sts\Фінальні результати обрані за 050426\SHA\Реслера\ShaОдинSeed\finalAnalysisReportRossShaOS.txt"),
#     (  r"C:\Users\sasha\Desktop\Диплом 25\ChaosCipher\DiplomAnalys\Експеремент Nist Sts\Фінальні результати обрані за 050426\SHA\Чуа\ShaОдинSeed\finalAnalysisReportChuaShaOS.txt"),
#     (r"C:\Users\sasha\Desktop\Диплом 25\ChaosCipher\DiplomAnalys\Експеремент Nist Sts\Фінальні результати обрані за 050426\SHA\Дуффінга\ShaОдинSeed\finalAnalysisReportDuffShaOS.txt"),
#     (r"C:\Users\sasha\Desktop\Диплом 25\ChaosCipher\DiplomAnalys\Експеремент Nist Sts\Фінальні результати обрані за 050426\SHA\ВанДерПоля\ShaОдинSeed\finalAnalysisReportVanPolShaOS.txt"),
#     (r"C:\Users\sasha\Desktop\Диплом 25\ChaosCipher\DiplomAnalys\Експеремент Nist Sts\Фінальні результати обрані за 050426\SHA\МаятникЩоЗмушує\ShaОдинSeed\finalAnalysisReportForcedShaOS.txt"),
# ]



SYSTEMS = [
    (r"C:\Users\sasha\Desktop\Диплом 25\ChaosCipher\DiplomAnalys\Експеремент Nist Sts\Фінальні результати обрані за 050426\SHA\Лоренца\ShaДекількаSeed\finalAnalysisReportLorenShaМS.txt"),
    (r"C:\Users\sasha\Desktop\Диплом 25\ChaosCipher\DiplomAnalys\Експеремент Nist Sts\Фінальні результати обрані за 050426\SHA\Реслера\ShaДекількаSeed\finalAnalysisReportRossShaМS.txt"),
    (r"C:\Users\sasha\Desktop\Диплом 25\ChaosCipher\DiplomAnalys\Експеремент Nist Sts\Фінальні результати обрані за 050426\SHA\Чуа\ShaДекількаSeed\finalAnalysisReportChuaShaМS.txt"),
    (r"C:\Users\sasha\Desktop\Диплом 25\ChaosCipher\DiplomAnalys\Експеремент Nist Sts\Фінальні результати обрані за 050426\SHA\Дуффінга\ShaДекількаSeed\finalAnalysisReportDuffShaМS.txt"),
    (r"C:\Users\sasha\Desktop\Диплом 25\ChaosCipher\DiplomAnalys\Експеремент Nist Sts\Фінальні результати обрані за 050426\SHA\ВанДерПоля\ShaДекількаSeed\finalAnalysisReportVanPolShaМS.txt"),
    (r"C:\Users\sasha\Desktop\Диплом 25\ChaosCipher\DiplomAnalys\Експеремент Nist Sts\Фінальні результати обрані за 050426\SHA\МаятникЩоЗмушує\ShaДекількаSeed\finalAnalysisReportForcedShaМS.txt"),
]

def create_statistic_file(input_file):
    INPUT_FILE = input_file
    file_path= INPUT_FILE.split('.')[0]
    OUTPUT_FILE = f"{file_path}.xlsx"

    rows = []
    random_excursions_threshold=0
    general_p_value_threshold = 0.945
    with open(INPUT_FILE, encoding="utf-8", errors="replace") as f:
        lines = f.readlines()
        print(lines)
        line_before_last_line = lines[-2]
        random_excursions_threshold = float(re.findall(r"[\d.]+", line_before_last_line)[-1])
        for line in lines:
            m = re.match(r'^((?:\s*\d+){10})\s+([\d.]+)\s+([\d.]+)\s*(\*)?\s+(.+?)\s*$', line)
            if m:
                counts_str, pvalue, proportion, flag, test = m.groups()
                counts = [int(x) for x in counts_str.split()]
                if len(counts) == 10:
                    rows.append({
                        "counts": counts,
                        "pvalue": float(pvalue),
                        "proportion": float(proportion),
                        "flagged": flag == "*",
                        "test": test.strip(),
                    })
            print(rows)



    print(f"Найдено рядків: {len(rows)}")

    HEADER_FILL = PatternFill("solid", start_color="1F4E79")
    ALT_FILL = PatternFill("solid", start_color="D6E4F0")
    WHITE_FILL = PatternFill("solid", start_color="FFFFFF")
    FLAG_FILL = PatternFill("solid", start_color="FFE0E0")
    PASS_FILL = PatternFill("solid", start_color="E2EFDA")
    FAIL_FILL = PatternFill("solid", start_color="FFDDC1")
    h_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    d_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    THRESHOLD = 0.001


    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NIST Results"
    ws.freeze_panes = "A2"

    headers = ["#", "C1", "C2", "C3", "C4", "C5", "C6", "C7", "C8", "C9", "C10",
               "P-VALUE", "PROPORTION", "PASS P_V?", "PASS PR?", "STATISTICAL TEST"]
    col_widths = [5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5,
                  12, 12, 9, 9, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = h_font
        c.fill = HEADER_FILL
        c.alignment = center
        c.border = brd
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    for i, r in enumerate(rows, 1):
        rn = i + 1
        bg = WHITE_FILL if i % 2 else ALT_FILL
        if r["flagged"]:
            bg = FLAG_FILL
        passed_pv = r["pvalue"] >= THRESHOLD
        if(r["test"]=="RandomExcursions" or r["test"]=="RandomExcursionsVariant"):
            passed_prop = r["proportion"] >= random_excursions_threshold
        else:
            passed_prop = r["proportion"] >= general_p_value_threshold

        vals = [i] + r["counts"] + [
            r["pvalue"],
            r["proportion"],
            "PASS" if passed_pv else "FAIL",
            "PASS" if passed_prop else "FAIL",
            r["test"]
        ]

        aligns = [center] * 15 + [left]

        for col, (val, aln) in enumerate(zip(vals, aligns), 1):
            c = ws.cell(row=rn, column=col, value=val)
            c.font = d_font
            c.fill = bg
            c.alignment = aln
            c.border = brd

        ws.cell(rn, 12).number_format = "0.000000"
        ws.cell(rn, 12).font = Font(name="Arial", size=10,
                                    color="375623" if passed_pv else "9C0006")
        ws.cell(rn, 13).number_format = "0.000"

        pc = ws.cell(rn, 14)
        pc.fill = PASS_FILL if passed_pv else FAIL_FILL
        pc.font = Font(name="Arial", size=10, bold=True,
                       color="375623" if passed_pv else "9C0006")
        pc.border = brd
        pc.alignment = center
        pc2 = ws.cell(rn, 15)
        pc2.fill = PASS_FILL if passed_prop else FAIL_FILL
        pc2.font = Font(
            name="Arial",
            size=10,
            bold=True,
            color="375623" if passed_prop else "9C0006"
        )
        pc2.border = brd
        pc2.alignment = center



    ws2 = wb.create_sheet("Summary by Test")
    ws2.freeze_panes = "A2"

    sh = ["STATISTICAL TEST", "Total",
          "Passed P_V", "Failed P_V", "Pass Rate P_V",
          "Passed PR", "Failed PR", "Pass Rate PR",
          "Avg P-VALUE", "Avg PROPORTION"]
    sw = [30, 9, 11, 11, 14, 11, 11, 14, 14, 16]

    for col, (h, w) in enumerate(zip(sh, sw), 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = h_font
        c.fill = HEADER_FILL
        c.alignment = center
        c.border = brd
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 22

    groups = defaultdict(list)
    for r in rows:
        groups[r["test"]].append(r)

    for i, (test, rlist) in enumerate(sorted(groups.items()), 1):
        total = len(rlist)

        # Pass/Fail по P-VALUE
        passed_pv = sum(1 for r in rlist if r["pvalue"] >= THRESHOLD)
        failed_pv  = total - passed_pv
        rate_pv    = passed_pv / total if total else 0

        # Pass/Fail по PROPORTION (с учётом типа теста)
        def prop_passed(r):
            thr = random_excursions_threshold \
                  if r["test"] in ("RandomExcursions", "RandomExcursionsVariant") \
                  else general_p_value_threshold
            return r["proportion"] >= thr

        passed_pr = sum(1 for r in rlist if prop_passed(r))
        failed_pr  = total - passed_pr
        rate_pr    = passed_pr / total if total else 0

        avg_p  = sum(r["pvalue"]     for r in rlist) / total
        avg_pr = sum(r["proportion"] for r in rlist) / total

        bg = WHITE_FILL if i % 2 else ALT_FILL
        vals = [test, total,
                passed_pv, failed_pv, rate_pv,
                passed_pr, failed_pr, rate_pr,
                avg_p, avg_pr]

        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=i + 1, column=col, value=val)
            c.font = d_font
            c.fill = bg
            c.border = brd
            c.alignment = left if col == 1 else center

        for rate_col, rate_val in [(5, rate_pv), (8, rate_pr)]:
            rc = ws2.cell(row=i + 1, column=rate_col)
            rc.number_format = "0.0%"
            rc.font = Font(name="Arial", size=10, bold=True,
                           color="375623" if rate_val == 1.0 else
                                 "9C0006" if rate_val < 0.8 else "7F6000")

        ws2.cell(i + 1, 9).number_format  = "0.000000"
        ws2.cell(i + 1, 10).number_format = "0.000"
        ws2.row_dimensions[i + 1].height = 18
    wb.save(OUTPUT_FILE)
    print(f"Готово! {len(rows)} рідків збережено → {OUTPUT_FILE}")


for path in SYSTEMS:
    create_statistic_file(path)
