"""
Benchmark for chaotic PRNG systems — one seed only.
Measures: Float mode (generate_bin_from_one_seed)
          SHA-256 mode (generate_bin_from_one_seed_SHA)

Output size per run (~6.875 MB):
  Float  : 1 718 751 * 4 bytes = 6 875 004 bytes
  SHA-256: 6 875 001 bytes
"""

import time, statistics, math
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from encrypt_alg.test import generator_factory
from generateFileForNist.create_bin import generate_bin_from_one_seed, generate_bin_from_one_seed_SHA

# ── plug in your imports ──────────────────────────────────────────────────────
# from your_module import generator_factory, extract_mantissa_bits
# from your_nist import generate_bin_from_one_seed, generate_bin_from_one_seed_SHA
# ─────────────────────────────────────────────────────────────────────────────

MB_FLOAT = 1_718_751 * 4    # bytes produced by Float mode
MB_SHA   = 6_875_001        # bytes produced by SHA-256 mode

SYSTEMS = [
    {
        "name":   "Lorenz",
        "type":   "lorenz",
        "params": {"logisticXLorenz": 0.01, "lorenzX": -20, "lorenzY": -20, "lorenzZ": 0},
    },
    {
        "name":   "Rossler",
        "type":   "rossler",
        "params": {"logisticXRossler": 0.01, "rosslerX": -15, "rosslerY": -15, "rosslerZ": 0},
    },
    {
        "name":   "Chua",
        "type":   "chua",
        "params": {"logisticXChua": 0.01, "chuaX": -3, "chuaY": -2, "chuaZ": -3},
    },
    {
        "name":   "Duffing",
        "type":   "duffing",
        "params": {"logisticXDuffing": 0.01, "duffingX": -2, "duffingY": -2, "duffingT": 0},
    },
    {
        "name":   "Van der Pol",
        "type":   "pol",
        "params": {"logisticXPol": 0.01, "polX": -3, "polY": -8, "polT": 0},
    },
    {
        "name":   "Forced Pendulum",
        "type":   "forced",
        "params": {"logisticXForced": 0.01, "forcedX": -math.pi, "forcedY": -6, "forcedT": 0},
    },
]

N_REPEATS = 3


# ── helpers ───────────────────────────────────────────────────────────────────

def time_it(fn):
    times = []
    for _ in range(N_REPEATS):
        t0 = time.perf_counter()
        fn()
        times.append(time.perf_counter() - t0)
    return statistics.mean(times)

def mbps(byte_count, seconds):
    return round(byte_count / seconds / 1_048_576, 3)


# ── benchmark ─────────────────────────────────────────────────────────────────

def run_all_benchmarks():
    rows = []
    for sys in SYSTEMS:
        t, p, label = sys["type"], sys["params"], sys["name"]
        print(f"\n── {label} ──────────────────────────────")

        # Float
        gen = generator_factory.create_generator(t, p, "chars")
        sec = time_it(lambda: generate_bin_from_one_seed(t, gen))
        speed = mbps(MB_FLOAT, sec)
        print(f"  Float  : {speed:>8.3f} MB/s  ({sec*1000:.1f} ms)")
        rows.append((label, "Float",   speed, round(sec * 1000, 1)))

        # SHA-256
        gen_sha = generator_factory.create_generator(t, p, "bits")
        sec = time_it(lambda: generate_bin_from_one_seed_SHA(t, gen_sha))
        speed = mbps(MB_SHA, sec)
        print(f"  SHA-256: {speed:>8.3f} MB/s  ({sec*1000:.1f} ms)")
        rows.append((label, "SHA-256", speed, round(sec * 1000, 1)))

    return rows


# ── Excel ─────────────────────────────────────────────────────────────────────

C = {
    "header": "1F3864",
    "sub":    "2E75B6",
    "float":  "D6E4F7",
    "sha":    "FFF2CC",
    "best":   "C6EFCE",
    "worst":  "FFC7CE",
}

def brd():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def hcell(ws, r, c, v, bg):
    cell = ws.cell(r, c, v)
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = brd()

def dcell(ws, r, c, v, bg=None, bold=False, fmt=None):
    cell = ws.cell(r, c, v)
    cell.font      = Font(name="Arial", size=10, bold=bold)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = brd()
    if bg:  cell.fill = PatternFill("solid", start_color=bg)
    if fmt: cell.number_format = fmt


def build_excel(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.freeze_panes = "A3"

    # title
    ws.merge_cells("A1:D1")
    t = ws.cell(1, 1, f"Chaotic PRNG — One Seed  |  ~6.875 MB/run  |  {N_REPEATS} repeats")
    t.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill      = PatternFill("solid", start_color=C["header"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    for ci, h in enumerate(["System", "Mode", "Speed (MB/s)", "Avg time (ms)"], 1):
        hcell(ws, 2, ci, h, C["sub"])
    ws.row_dimensions[2].height = 24

    best_speed  = max(r[2] for r in rows)
    worst_speed = min(r[2] for r in rows)

    for ri, (name, mode, speed, ms) in enumerate(rows, start=3):
        row_bg  = C["float"] if mode == "Float" else C["sha"]
        cell_bg = C["best"] if speed == best_speed else (
                  C["worst"] if speed == worst_speed else row_bg)
        bold = speed in (best_speed, worst_speed)

        dcell(ws, ri, 1, name,  bg=cell_bg, bold=bold)
        dcell(ws, ri, 2, mode,  bg=cell_bg, bold=bold)
        dcell(ws, ri, 3, speed, bg=cell_bg, bold=bold, fmt="0.000")
        dcell(ws, ri, 4, ms,    bg=cell_bg, bold=bold, fmt="0.0")
        ws.row_dimensions[ri].height = 17

    for ci, w in enumerate([18, 10, 14, 14], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    lr = len(rows) + 4
    ws.cell(lr,   1, "Green = fastest  |  Red = slowest").font = Font(name="Arial", size=9, italic=True)
    ws.cell(lr+1, 1, "Blue = Float rows  |  Yellow = SHA-256 rows").font = Font(name="Arial", size=9, italic=True)

    wb.save(path)
    print(f"\nSaved -> {path}")


# ── entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    rows = run_all_benchmarks()
    build_excel(rows, r"C:\Users\sasha\Desktop\Диплом 25\ChaosCipher\DiplomAnalys\Експеремент Nist Sts\Фінальні результати обрані за 050426\benchmark_results.xlsx")